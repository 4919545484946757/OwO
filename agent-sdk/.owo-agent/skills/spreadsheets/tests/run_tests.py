# -*- coding: utf-8 -*-
"""spreadsheets 技能契约测试：xlsx 生成 / 公式 / CSV 往返（>=3 个端到端用例）。"""
import csv
import os
import shutil
import tempfile

from openpyxl import Workbook, load_workbook


def case1(tmp: str) -> None:
    """生成 xlsx 并断言表头与单元格值。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销量"
    sheet.append(["名称", "数量", "单价"])
    sheet.append(["苹果", 3, 2.5])
    sheet.append(["香蕉", 5, 1.2])
    path = os.path.join(tmp, "sales.xlsx")
    workbook.save(path)
    loaded = load_workbook(path)
    rows = list(loaded.active.iter_rows(values_only=True))
    assert rows[0] == ("名称", "数量", "单价")
    assert rows[1] == ("苹果", 3, 2.5)
    assert rows[2][0] == "香蕉"
    print("case1 ok: 生成 xlsx 并断言单元格")


def case2(tmp: str) -> None:
    """公式单元格写入与往返保留。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["项目", "金额"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])
    sheet.append(["C", 30])
    sheet["B5"] = "=SUM(B2:B4)"
    sheet["B6"] = "=AVERAGE(B2:B4)"
    path = os.path.join(tmp, "formula.xlsx")
    workbook.save(path)
    loaded = load_workbook(path)
    assert loaded.active["B5"].value == "=SUM(B2:B4)"
    assert loaded.active["B6"].value == "=AVERAGE(B2:B4)"
    print("case2 ok: 公式单元格往返")


def case3(tmp: str) -> None:
    """CSV -> xlsx -> CSV 往返，数据无丢失。"""
    csv_path = os.path.join(tmp, "input.csv")
    xlsx_path = os.path.join(tmp, "converted.xlsx")
    csv_out = os.path.join(tmp, "output.csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "name", "score"])
        writer.writerow([1, "张三", 88])
        writer.writerow([2, "李四", 92])

    workbook = Workbook()
    sheet = workbook.active
    with open(csv_path, encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            sheet.append(row)
    workbook.save(xlsx_path)

    loaded = load_workbook(xlsx_path)
    with open(csv_out, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in loaded.active.iter_rows(values_only=True):
            writer.writerow(row)
    with open(csv_out, encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows == [["id", "name", "score"], ["1", "张三", "88"], ["2", "李四", "92"]]
    print("case3 ok: CSV 往返无丢失")


def main() -> None:
    tmp = tempfile.mkdtemp(prefix="owo-skill-spreadsheets-")
    try:
        case1(tmp)
        case2(tmp)
        case3(tmp)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    main()
